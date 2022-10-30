import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    return(sheet.max_row)

def getColumnCount(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    return(sheet.max_column)

def readData(file,SheetName,rownumber,columnnumber):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    return sheet.cell(rownumber,columnnumber).value

def writeData(file,SheetName,rownumber,columnnumber,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    sheet.cell(rownumber,columnnumber).value=data
    workbook.save(file)

def FillGreenColor(file,SheetName,rownumber,columnnumber):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    greenFill=PatternFill(start_color='60b212',
                          end_color='60b212',
                          fill_type='solid')
    sheet.cell(rownumber,columnnumber).fill=greenFill
    workbook.save(file)

def FillRedColor(file,SheetName,rownumber,columnnumber):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    redFill=PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(rownumber,columnnumber).fill=redFill
    workbook.save(file)